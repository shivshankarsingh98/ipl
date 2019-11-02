import openpyxl
import os
import django

os.environ.setdefault('DJANGO_SETTINGS_MODULE','ipl.settings')
django.setup()

from iplapp.models import*


def populate_matches():
    source = "matches.xlsx"
    src_workbook = openpyxl.load_workbook(source)
    for row in src_workbook.worksheets[0].iter_rows(values_only=False):
        words = []
        for cell in row:
            words.append(cell._value)
        if words[0] != 'id' and words[0]!=None:
            p = matches(
                id=words[0],
                season=words[1],
                city=words[2],
                date=words[3],
                team1=words[4],
                team2=words[5],
                toss_winner=words[6],
                toss_decision=words[7],
                result=words[8],
                dl_applied=words[9],
                winner=words[10],
                win_by_run=words[11],
                win_by_wicket=words[12],
                player_of_match=words[13],
                venue=words[14],
                umpire1=words[15],
                umpire2=words[16],
                umpire3=words[17],

            )
            p.save()


def populate_deliveries():
    source = "deliveries.xlsx"
    src_workbook = openpyxl.load_workbook(source)
    for row in src_workbook.worksheets[0].iter_rows(values_only=False):
        try:
            words = []
            for cell in row:
                words.append(cell._value)
            if words[0] != 'match_id' and words[0] != None:
                p = deliveries(
                    match_id= matches.objects.get(id=words[0]),
                inning = words[1],
                batting_team = words[2],
                bowling_team = words[3],
                over = words[4],
                ball = words[5],
                batsman = words[6],
                non_striker = words[7],
                bowler = words[8],

                is_super_over = words[9],
                wide_runs = words[10],
                bye_runs = words[11],
                legbye_runs = words[12],
                noball_runs = words[13],
                penalty_runs = words[14],
                batsman_runs = words[15],
                extra_runs = words[16],
                total_runs = words[17],

                player_dismissed = words[18],
                dismissal_kind = words[19],
                fielder = words[20],




                )
                p.save()
        except Exception as exp:
            print("Failed row:", row)
            print("Reason:", exp)



populate_deliveries()
#populate_matches()